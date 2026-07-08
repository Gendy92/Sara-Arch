#!/usr/bin/env python3
"""Convert ACTION_PLAN.md to an Excel workbook with one sheet per section."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SPEC_PATH = Path('ACTION_PLAN.md')
OUT_PATH = Path('Sara-Arch_Action_Plan.xlsx')

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=12)
SUB_FILL = PatternFill('solid', fgColor='D9E1F2')
SUB_FONT = Font(bold=True, size=11)
TABLE_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
TABLE_HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BOLD_FONT = Font(bold=True)
RTL_ALIGN = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)

def strip_bold_markers(text):
    return re.sub(r'\*\*([^\*]+)\*\*', r'\1', text)

def parse_blocks(lines):
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.rstrip()
        if not stripped:
            i += 1
            continue
        if stripped.startswith('### '):
            yield ('heading', stripped.lstrip('#').strip())
            i += 1
            continue
        if stripped.startswith('|'):
            rows = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                row = [cell.strip() for cell in lines[i].strip().strip('|').split('|')]
                if not all(re.match(r'^[\-\:\s]+$', c) for c in row):
                    rows.append(row)
                i += 1
            if rows:
                yield ('table', rows)
            continue
        yield ('para', stripped)
        i += 1

def split_sections(text):
    sections = []
    pre = []
    current_title = None
    current_lines = []
    for line in text.splitlines():
        m = re.match(r'^##\s+(.*)$', line)
        if m:
            if current_title is not None:
                sections.append((current_title, current_lines))
            else:
                pre = current_lines
            current_title = m.group(1).strip()
            current_lines = []
        else:
            current_lines.append(line)
    if current_title is not None:
        sections.append((current_title, current_lines))
    else:
        pre = current_lines
    return pre, sections

def write_sheet(ws, title, blocks):
    ws.sheet_view.rightToLeft = True
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(1, 1)
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = RTL_ALIGN

    row_idx = 2
    for kind, data in blocks:
        if kind == 'heading':
            ws.append([data])
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
            cell = ws.cell(row_idx, 1)
            cell.font = SUB_FONT
            cell.fill = SUB_FILL
            cell.alignment = RTL_ALIGN
            row_idx += 1
        elif kind == 'para':
            txt = strip_bold_markers(data)
            ws.append([txt])
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
            cell = ws.cell(row_idx, 1)
            cell.alignment = RTL_ALIGN
            if '**' in data:
                cell.font = BOLD_FONT
            row_idx += 1
        elif kind == 'table':
            rows = data
            max_cols = max(len(r) for r in rows)
            header_row = row_idx
            for r in rows:
                r += [''] * (max_cols - len(r))
                ws.append(r)
                for col in range(1, max_cols + 1):
                    cell = ws.cell(row_idx, col)
                    cell.alignment = RTL_ALIGN
                    cell.border = border
                    if row_idx == header_row:
                        cell.fill = TABLE_HEADER_FILL
                        cell.font = TABLE_HEADER_FONT
                row_idx += 1
            ws.append([''])
            row_idx += 1

    for col in range(1, 9):
        max_len = 0
        for r in range(1, row_idx):
            cell = ws.cell(r, col)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 2, 80)
        ws.column_dimensions[get_column_letter(col)].width = max(width, 15)

    ws.freeze_panes = 'A2'

def main():
    text = SPEC_PATH.read_text(encoding='utf-8')
    pre, sections = split_sections(text)

    wb = Workbook()
    wb.remove(wb.active)

    toc = wb.create_sheet('TOC')
    toc.sheet_view.rightToLeft = True
    toc.append(['جدول المحتويات / Table of Contents'])
    toc.merge_cells('A1:B1')
    toc.cell(1,1).font = HEADER_FONT
    toc.cell(1,1).fill = HEADER_FILL
    toc.cell(1,1).alignment = RTL_ALIGN
    toc.append(['Sheet', 'Section'])
    for c in toc[2]:
        c.font = TABLE_HEADER_FONT
        c.fill = TABLE_HEADER_FILL
        c.alignment = RTL_ALIGN

    meta = wb.create_sheet('Metadata')
    meta.sheet_view.rightToLeft = True
    meta_blocks = list(parse_blocks(pre))
    write_sheet(meta, 'Metadata', meta_blocks)

    for raw_title, lines in sections:
        sheet_title = re.sub(r'[\\/*?:\[\]]', '-', raw_title)[:31]
        blocks = list(parse_blocks(lines))
        ws = wb.create_sheet(sheet_title)
        write_sheet(ws, raw_title, blocks)
        toc.append([sheet_title, raw_title])

    for row in toc.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = RTL_ALIGN
    toc.column_dimensions['A'].width = 35
    toc.column_dimensions['B'].width = 70

    wb.save(OUT_PATH)
    print(f'Saved {OUT_PATH} with {len(sections)} sections + TOC + Metadata')

if __name__ == '__main__':
    main()
